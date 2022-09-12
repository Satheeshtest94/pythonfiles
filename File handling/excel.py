#Writing in workbook
"""
from openpyxl import Workbook

# Creating the workbook
wb = Workbook()
# Creating active worksheet
ws = wb.active
# Changing the title of worksheet
ws.title = 'No1'
ws1 = wb.create_sheet("No2",1)
ws2 = wb.create_sheet("No5",-1)
# Writing data to the worksheet
ws['A1'] = "I am here"
# Saving the workbook
wb.save("4th Sep.Xlsx")
# Closed the workbook
wb.close()

import os

print(os.getcwd())

"""

#Reading from workbook

import openpyxl
from openpyxl import Workbook

#Loding the workbook
wb = openpyxl.load_workbook("4th Sep.Xlsx")
ws = wb.active
#print(wb.sheetnames)
#print(ws)

#print(ws.values)

for row in range(1,5):
  for k in row:
   print(k,end = " ")
  print()





#d = ws.cell(row = 1,column = 1)
#print(d.value)

#iter_rows # iter through rows
#item_column #iter through colo


#for x in range(1,5):
    #for y in range(1,5):
        #print(x,y,ws.cell(row = x,column = y))



# To find max rows & column

#print(ws.max_row)
#print(ws.max_column)

"""

for row in ws.iter_rows(min_row = 1,max_col = 5,max_row = 2):
    for ws.values in row:
     print(ws.values)

"""

   
"""

#print(ws['A1'].value)
"""









