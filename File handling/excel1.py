import openpyxl

wb = openpyxl.load_workbook()
ws = wb.active
ws.title = "Started new"
ws['A1'] = "own try"
wb.save("Mine.xlsx")
wb.close()

wb = openpyxl.load_workbook("Mine.xlsx")
m = ws.cell(row = 1,column =1)
print(m.value)
