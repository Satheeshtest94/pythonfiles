from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws['A5'] = "Hello"

v1 = ws['A10']

v1.value = 500

#j = input("Enter the data:")

#ws.cell(row= 1,column = 5).value = j

#wb.save("C:\\Users\\Sathe\\PycharmProjects\\pythonProject1\\File handling\\4th Sep.xlsx")



for i in range(1,3):
    for j in range(1,3):
     k = input("Enter the data:")
    ws.cell(row=i, column=j).value = k


wb.save("C:\\Users\\Sathe\\PycharmProjects\\pythonProject1\\File handling\\4th Sep.xlsx")




