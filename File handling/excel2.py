import openpyxl

path = "/File handling/4th Sep.Xlsx"
wb = openpyxl.load_workbook(path)
#print(wb)
ws = wb.active
#d = ws.cell(row = 1, column =1)
#print(d.value)

for rows in range(1,4):
    for col in range(1,4):
     m = ws.cell(row = rows,column = col)
     print(m.value,end = " ")
    print()


