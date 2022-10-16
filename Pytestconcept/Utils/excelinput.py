

import openpyxl
from openpyxl import Workbook

# print(new_sheet.max_row)
class datainput():

    @staticmethod
    def handling():

            Dict = {}

            path = "C:\\Users\\Sathe\\PycharmProjects\\pythonProject1\\Pythonclass\\Pytestconcept\\Input file\\input file.xlsx"

            new_book = openpyxl.load_workbook(path)

            Sheet = new_book.active
            # data  = new_sheet.cell(row =1 ,column=1)
            # print(data.value)
            max_row = Sheet.max_row
            max_col = Sheet.max_column
            # # print(new_sheet.max_column)
            # for row in range(1,(max_row)+1):
            #     for column in range(1,(max_col)+1):
            #         cell_value = new_sheet.cell(row = row,column=column).value

            for i in range(2, max_row + 1):
                for j in range(1, max_col + 1):
                   Dict[(Sheet.cell(row=1, column=j).value)] = Sheet.cell(row=i, column=j).value
                #print(dict)
            return Dict

obj = datainput()
obj.handling()